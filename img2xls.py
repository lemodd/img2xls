from PIL import Image
import numpy

from openpyxl.styles import PatternFill,Border,Side
from openpyxl import Workbook

import sys

if len(sys.argv)>1:
    img_file = sys.argv[1]
else:
    img_file = 'czj.jpg'

print(img_file)

wb = Workbook()
ws1 = wb.active
ws1.title = 's1'                

border = Border(left=Side(style='thin',color='FF000000'),
                right=Side(style='thin',color='FF000000'),
                top=Side(style='thin',color='FF000000'),
                bottom=Side(style='thin',color='FF000000'),
                diagonal=Side(style='thin',color='FF000000'),
                diagonal_direction=0,
                outline=Side(style='thin',color='FF000000'),
                vertical=Side(style='thin',color='FF000000'),
                horizontal=Side(style='thin',color='FF000000'))

image = Image.open(img_file)
array = numpy.array(image)
color_hex =''
row_count = len(array)
col_count = len(array[0])

#此函数来自https://jalena.bcsytv.com/archives/2113
def convent_column_to_char(column):

    if not isinstance(column, int):
            return column
    tStr = str()
    while column != 0:
        res = column % 26
        if res == 0:
                res = 26
                column -= 26
        tStr = chr(ord('A') + res - 1) + tStr
        column = column // 26
    return tStr


print('正在转换，请稍候。。。')

for i in range(row_count):
    for j in range(col_count):
        pixel = array[i][j]
        color_hex = ''
        for p in pixel:
            temp = hex(p)[2:].upper()
            if len(temp) == 1:
                temp = '0' + temp
            color_hex += temp
            
        fill = PatternFill(color_hex, color_hex, fill_type = 'solid')
        ws1.cell(i+1,j+1).fill = fill

col_count = len(array[0])

for i in range(col_count):
    i += 1
    cname = convent_column_to_char(i)

    ws1.column_dimensions[cname].width = 2

    for j in range(row_count):
        j += 1
        cell_num = cname+str(j)
        ws1[cell_num].border=border
        ws1.row_dimensions[j].height = 12


xls_file = img_file.split('.')[0]+'.xlsx'
wb.save(xls_file)

input('完成，回车继续.')
